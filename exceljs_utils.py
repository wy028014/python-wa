import openpyxl

class ExcelJS:
    def read_file(self, file_path):
        data = []
        if '轨迹查询.xlsx' in file_path:
            workbook = openpyxl.load_workbook(file_path)
            worksheet = workbook.active
            for row in worksheet.iter_rows(values_only=True):
                data.append({
                    '姓名': row[0],
                    '证件类型': row[1],
                    '证件编号': row[2],
                    '乘车日期': row[3],
                    '乘车时间': row[4],
                    '车次': row[5],
                    '发站': row[6],
                    '到站': row[7],
                    '车厢号': row[8],
                    '席别': row[9],
                    '座位号': row[10],
                    '票价': row[11]
                })
        elif '站站查询.xlsx' in file_path:
            workbook = openpyxl.load_workbook(file_path)
            worksheet = workbook.active
            for row in worksheet.iter_rows(values_only=True):
                data.append({
                    '姓名': row[0],
                    '证件类型': row[1],
                    '证件编号': row[2],
                    '乘车日期': row[3],
                    '乘车时间': row[4],
                    '票号': row[5],
                    '车次': row[6],
                    '发站': row[7],
                    '到站': row[8],
                    '车厢号': row[9],
                    '席别': row[10],
                    '座位号': row[11],
                    '票种': row[12],
                    '票价': row[13],
                    '售票处': row[14],
                    '窗口': row[15],
                    '操作员': row[16],
                    '售票时间': row[17]
                })
        if data:
            data.pop(0)
        return data

exceljs = ExcelJS()