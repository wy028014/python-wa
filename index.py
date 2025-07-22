# -*- coding: utf-8 -*-
import asyncio
import threading
import traceback
import logging
import tempfile
import os
import uuid
import random
import re
import json
import openpyxl
import pyautogui
import pygetwindow as gw
from functools import wraps
from datetime import datetime
from flask import Flask, request, jsonify, make_response
from playwright.async_api import async_playwright

app = Flask(__name__)
app.config['JSON_AS_ASCII'] = False
logger = logging.getLogger('AutomationServer')

# 配置日志
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(name)s - %(levelname)s - %(message)s')

# 线程局部存储
_thread_local = threading.local()

# ====================== Excel 处理工具 ======================
class ExcelProcessor:
    COLUMN_MAPPING = {
        'glcx': {
            '业务类型': ['业务类型', '业务类别'],
            '姓名': ['姓名', '乘客姓名'],
            '证件类型': ['证件类型', '证件类别'],
            '证件编号': ['证件编号', '证件号码', '身份证号'],
            '乘车日期': ['乘车日期', '出行日期'],
            '乘车时间': ['乘车时间', '出发时间'],
            '车次': ['车次', '列车号'],
            '发站': ['发站', '出发站'],
            '到站': ['到站', '到达站'],
            '车厢号': ['车厢号', '车厢'],
            '席别': ['席别', '座位类型'],
            '座位号': ['座位号', '座位'],
            '票价': ['票价', '金额']
        },
        'zzcx': {
            '姓名': ['姓名', '乘客姓名'],
            '证件类型': ['证件类型', '证件类别'],
            '证件编号': ['证件编号', '证件号码', '身份证号'],
            '乘车日期': ['乘车日期', '出行日期'],
            '乘车时间': ['乘车时间', '出发时间'],
            '票号': ['票号', '票据号'],
            '车次': ['车次', '列车号'],
            '发站': ['发站', '出发站'],
            '到站': ['到站', '到达站'],
            '车厢号': ['车厢号', '车厢'],
            '席别': ['席别', '座位类型'],
            '座位号': ['座位号', '座位'],
            '票种': ['票种', '票据类型'],
            '票价': ['票价', '金额'],
            '售票处': ['售票处', '售票点'],
            '窗口': ['窗口', '柜台'],
            '操作员': ['操作员', '售票员'],
            '售票时间': ['售票时间', '出票时间']
        },
        'plgjcx': {
            '姓名': ['姓名', '乘客姓名'],
            '证件类型': ['证件类型', '证件类别'],
            '证件编号': ['证件编号', '证件号码', '身份证号'],
            '乘车日期': ['乘车日期', '出行日期'],
            '乘车时间': ['乘车时间', '出发时间'],
            '车次': ['车次', '列车号'],
            '发站': ['发站', '出发站'],
            '到站': ['到站', '到达站'],
            '车厢号': ['车厢号', '车厢'],
            '席别': ['席别', '座位类型'],
            '座位号': ['座位号', '座位'],
            '票价': ['票价', '金额']
        }
    }
    
    DATE_FORMATS = [
        '%Y-%m-%d',
        '%Y/%m/%d',
        '%Y%m%d',
        '%Y-%m-%d %H:%M:%S',
        '%Y/%m/%d %H:%M:%S',
        '%Y%m%d %H:%M:%S',
        '%Y-%m-%d %H:%M',
        '%Y/%m/%d %H:%M',
        '%Y%m%d %H:%M'
    ]
    
    def _clean_value(self, value):
        """清理和转换Excel值"""
        if value is None:
            return None
            
        if isinstance(value, str):
            # 移除不可见字符和多余空格
            value = re.sub(r'\s+', ' ', value).strip()
            
            # 尝试转换日期格式
            for fmt in self.DATE_FORMATS:
                try:
                    dt = datetime.strptime(value, fmt)
                    if ' ' in fmt:
                        return dt.isoformat(sep=' ')
                    else:
                        return dt.date().isoformat()
                except ValueError:
                    continue
        
        # 如果是datetime对象，转换为字符串
        if isinstance(value, datetime):
            return value.isoformat(sep=' ')
            
        return value

    def read_file(self, file_path, query_type):
        """读取Excel文件并返回结构化数据"""
        if not os.path.exists(file_path):
            logger.error(f"文件不存在: {file_path}")
            return []
            
        try:
            return self._read_excel(file_path, query_type)
        except Exception as e:
            logger.error(f"读取文件 {file_path} 时出错: {str(e)}")
            logger.debug(traceback.format_exc())
            return []

    def _read_excel(self, file_path, query_type):
        """读取特定类型的Excel文件"""
        if query_type not in self.COLUMN_MAPPING:
            logger.error(f"未知的查询类型: {query_type}")
            return []
            
        # 获取列映射
        column_map = self.COLUMN_MAPPING[query_type]
        data = []
        
        try:
            workbook = openpyxl.load_workbook(file_path, data_only=True)
            worksheet = workbook.active
            
            # 记录读取进度
            total_rows = worksheet.max_row
            logger.info(f"开始读取 {file_path} (类型: {query_type}), 共 {total_rows} 行")
            
            # 读取标题行，构建列索引映射
            header_row = next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True))
            column_index_map = {}
            
            # 创建标题到索引的映射
            for idx, header in enumerate(header_row):
                if header is None:
                    continue
                    
                # 尝试匹配每个标准列名
                for col_name, aliases in column_map.items():
                    if header in aliases:
                        column_index_map[col_name] = idx
                        break
            
            # 检查是否所有必要列都存在
            missing_columns = set(column_map.keys()) - set(column_index_map.keys())
            if missing_columns:
                logger.warning(f"缺少列: {', '.join(missing_columns)}，将使用默认映射")
                # 回退到默认索引（如果可能）
                for col_name in missing_columns:
                    if col_name in self.COLUMN_MAPPING.get(query_type, {}):
                        # 尝试使用第一个别名作为默认
                        column_index_map[col_name] = self.COLUMN_MAPPING[query_type][col_name][0]
            
            # 跳过标题行，从第2行开始
            processed_rows = 0
            for row_idx, row in enumerate(worksheet.iter_rows(min_row=2, values_only=True), start=2):
                # 跳过空行
                if all(cell is None or cell == '' for cell in row):
                    continue
                    
                item = {}
                valid_row = False
                
                for col_name, col_idx in column_index_map.items():
                    try:
                        if col_idx < len(row):
                            value = row[col_idx]
                        else:
                            value = None
                            
                        cleaned_value = self._clean_value(value)
                        item[col_name] = cleaned_value
                        
                        # 如果至少有一个列有值，则认为是有效行
                        if cleaned_value is not None and cleaned_value != '':
                            valid_row = True
                    except Exception as e:
                        logger.warning(f"处理行 {row_idx} 列 {col_name} 时出错: {str(e)}")
                        item[col_name] = None
                
                # 只添加有实际数据的行
                if valid_row:
                    data.append(item)
                    processed_rows += 1
                    
                    # 每处理100行记录一次进度
                    if processed_rows % 100 == 0:
                        logger.info(f"已处理 {processed_rows}/{total_rows - 1} 行")
            
            logger.info(f"成功读取 {processed_rows} 行数据")
            return data
        except Exception as e:
            logger.error(f"读取 {query_type} Excel 文件时出错: {str(e)}")
            logger.debug(traceback.format_exc())
            return []
        finally:
            # 尝试关闭工作簿
            try:
                workbook.close()
            except:
                pass

# ====================== Playwright 自动化工具 ======================
class PlaywrightAutomator:
    def __init__(self):
        self.playwright = None
        self.browser = None
        self.timeout = 60000
        self.data_dir = tempfile.mkdtemp(prefix="pw_data_")
        self._initialized = False
        self.lock = asyncio.Lock()

    async def initialize(self):
        """异步初始化Playwright环境"""
        async with self.lock:
            if self._initialized:
                logger.info("浏览器环境已初始化，跳过重复初始化")
                return
                
            try:
                logger.info("正在初始化Playwright环境...")
                self.playwright = await async_playwright().start()
                self.browser = await self.playwright.chromium.launch(
                    headless=False,
                    args=[
                        '--start-maximized',
                        '--disable-web-security',
                        '--ignore-certificate-errors',
                        '--allow-insecure-localhost'
                    ],
                    slow_mo=1000
                )
                self._initialized = True
                logger.info("Playwright环境初始化成功")
            except Exception as e:
                logger.error(f"初始化失败: {e}")
                traceback.print_exc()
                self._initialized = False
                raise RuntimeError(f"浏览器实例初始化失败: {str(e)}")
    
    def is_initialized(self):
        return self._initialized
    
    async def close(self):
        """关闭所有资源"""
        async with self.lock:
            try:
                if self.browser:
                    await self.browser.close()
                if self.playwright:
                    await self.playwright.stop()
                self._initialized = False
                logger.info("Playwright资源已关闭")
            except Exception as e:
                logger.error(f"关闭资源时出错: {e}")
                return False
            return True

    async def _handle_certificate_popup(self):
        """处理Windows数字证书弹窗"""
        try:
            await asyncio.sleep(2)
            cert_windows = gw.getWindowsWithTitle("数字证书")
            if cert_windows:
                window = cert_windows[0]
                window.activate()
                pyautogui.write("公共密码")  # 替换为实际密码
                pyautogui.press("enter")
                await asyncio.sleep(1)
                return True
            return False
        except Exception as e:
            logger.error(f"处理证书弹窗失败: {e}")
            return False

    async def test_login(self):
        """测试登录流程并返回网站标题"""
        if not self._initialized:
            await self.initialize()
            
        try:
            # 为测试创建独立的上下文和页面
            context = await self.browser.new_context(
                ignore_https_errors=True,
                accept_downloads=True,
                viewport={'width': 1920, 'height': 1080}
            )
            page = await context.new_page()
            page.on("dialog", lambda dialog: dialog.accept())
            page.on("certificateerror", lambda error: error.continue_())
            
            # 导航到登录页
            await page.goto('https://10.3.2.201:9943/rntibp/login.html', timeout=self.timeout)
            
            # 处理证书错误页面
            if await page.is_visible('text="此网站的安全证书存在问题"', timeout=5000):
                await page.click('text=高级')
                await page.click('text=继续前往(不安全)')
                await self._handle_certificate_popup()
            
            # 登录处理
            if await page.is_visible('#loginBtn', timeout=5000):
                await page.fill('input[type="password"]', '111111')
                await page.click('#loginBtn')
                await page.wait_for_load_state('networkidle')
            
            # 验证登录成功
            await page.wait_for_selector('.dashboard', timeout=self.timeout)
            logger.info("登录成功")
            
            # 获取网站标题
            title = await page.title()
            
            return title
        except Exception as e:
            logger.error(f"登录测试失败: {e}")
            return f"登录测试失败: {str(e)}"
        finally:
            # 清理资源
            if page:
                await page.close()
            if context:
                await context.close()

    async def _ensure_login(self, page):
        """确保用户已登录"""
        try:
            # 导航到登录页
            await page.goto('https://10.3.2.201:9943/rntibp/login.html', timeout=self.timeout)
            
            # 处理证书错误页面
            if await page.is_visible('text="此网站的安全证书存在问题"', timeout=5000):
                await page.click('text=高级')
                await page.click('text=继续前往(不安全)')
                await self._handle_certificate_popup()
            
            # 登录处理
            if await page.is_visible('#loginBtn', timeout=5000):
                await page.fill('input[type="password"]', '111111')
                await page.click('#loginBtn')
                await page.wait_for_load_state('networkidle')
            
            # 验证登录成功
            await page.wait_for_selector('.dashboard', timeout=self.timeout)
            logger.info("登录成功")
            return True
        except Exception as e:
            logger.error(f"登录失败: {e}")
            return False

    async def _prepare_form(self, page, query_type, params):
        """安全操作表单元素"""
        try:
            # 设置背景色
            await page.locator('.main-padding').evaluate(
                "node => node.style.backgroundColor = '#f0f0f0'"
            )
            
            if query_type == 'glcx':
                # 移除只读属性
                await page.locator('input[name="startDate"]').evaluate(
                    "node => node.removeAttribute('readonly')"
                )
                await page.locator('input[name="endDate"]').evaluate(
                    "node => node.removeAttribute('readonly')"
                )
                
                # 设置值
                await page.fill('input[name="startDate"]', params['date_start'])
                await page.fill('input[name="endDate"]', params['date_end'])
                await page.fill('#idNo', params['id_no'])
                
            elif query_type == 'zzcx':
                await page.locator('input[name="trainDate"]').evaluate(
                    "node => node.removeAttribute('readonly')"
                )
                await page.fill('input[name="trainDate"]', params['train_date'])
                await page.fill('#boardTrainCode', params['train_code'])
                await page.fill('#fromStation', params['from_station'])
                
                # 可选字段处理
                if params.get('to_station'):
                    await page.fill('#toStation', params['to_station'])
            
            elif query_type == 'plgjcx':
                await page.locator('input[name="startDate"]').evaluate(
                    "node => node.removeAttribute('readonly')"
                )
                await page.locator('input[name="endDate"]').evaluate(
                    "node => node.removeAttribute('readonly')"
                )
                await page.fill('input[name="startDate"]', params['date_start'])
                await page.fill('input[name="endDate"]', params['date_end'])
                
            return True
        except Exception as e:
            logger.error(f"表单准备失败: {e}")
            return False

    async def _download_excel(self, page, filename_prefix):
        """安全处理文件下载"""
        save_path = None
        try:
            # 创建唯一文件名
            file_name = f"{filename_prefix}_{uuid.uuid4().hex}.xlsx"
            save_path = os.path.join(self.data_dir, file_name)
            
            # 等待下载开始
            async with page.expect_download(timeout=self.timeout) as download_info:
                await page.click('#download')
                await asyncio.sleep(1)  # 确保点击生效
            
            download = await download_info.value
            
            # 保存文件
            await download.save_as(save_path)
            
            # 处理Excel
            processor = ExcelProcessor()
            result = processor.read_file(save_path, filename_prefix.split('_')[0])
            return result
        except Exception as e:
            logger.error(f"下载失败: {e}")
            return []
        finally:
            # 确保清理文件
            if save_path and os.path.exists(save_path):
                try:
                    os.unlink(save_path)
                except:
                    pass

    async def _perform_query(self, page, query_type, params):
        """执行查询操作"""
        save_path = None
        try:
            # 准备表单
            if not await self._prepare_form(page, query_type, params):
                return []
            
            # 定义每个路由的点击序列
            click_sequences = {
                'glcx': [
                    '#idNo',  # 点击身份证号输入框
                    '#startDate',  # 点击开始日期
                    '#endDate',  # 点击结束日期
                    '#queryBtn',  # 查询按钮
                    '#confirmBtn',  # 确认按钮（如果需要）
                    '#download'  # 下载按钮
                ],
                'zzcx': [
                    '#trainDate',
                    '#boardTrainCode',
                    '#fromStation',
                    '#toStation',
                    '#queryBtn',
                    '#confirmBtn',
                    '#download'
                ],
                'plgjcx': [
                    '#startDate',
                    '#endDate',
                    '#queryBtn',
                    '#confirmBtn',
                    '#download'
                ]
            }
            
            # 批量查询特殊处理 - 上传文件
            if query_type == 'plgjcx':
                # 生成ID列表文件
                txt_path = os.path.join(self.data_dir, f'ids_{uuid.uuid4().hex}.txt')
                with open(txt_path, 'w', encoding='utf-8') as f:
                    for id_no in params['id_no_list']:
                        f.write(id_no + '\n')
                
                try:
                    # 上传文件
                    file_input = await page.wait_for_selector('input[type=file]', timeout=10000)
                    await file_input.set_input_files(txt_path)
                    
                    # 确保文件上传按钮在点击序列中
                    if '#uploadBtn' not in click_sequences['plgjcx']:
                        click_sequences['plgjcx'].insert(2, '#uploadBtn')
                finally:
                    if os.path.exists(txt_path):
                        os.unlink(txt_path)
            
            # 创建下载事件监听
            async with page.expect_download(timeout=self.timeout) as download_info:
                # 执行路由特定的点击序列
                for selector in click_sequences[query_type]:
                    try:
                        # 添加随机延迟模拟人工操作
                        await asyncio.sleep(random.uniform(0.2, 0.5))
                        
                        # 更稳健的点击实现
                        element = await page.wait_for_selector(selector, timeout=5000)
                        await element.scroll_into_view_if_needed()
                        box = await element.bounding_box()
                        
                        # 点击元素中心位置
                        await page.mouse.click(
                            box['x'] + box['width']/2,
                            box['y'] + box['height']/2
                        )
                    except Exception as e:
                        logger.warning(f"点击元素 {selector} 失败: {e}")
                        # 尝试备用点击方法
                        await page.click(selector, timeout=2000, force=True)
                    
                    # 等待操作完成（根据不同操作调整等待时间）
                    if selector == '#uploadBtn':
                        await page.wait_for_selector('.upload-success', timeout=30000)
                    elif selector == '#download':
                        await asyncio.sleep(1)  # 确保下载开始
                    else:
                        await page.wait_for_timeout(300)  # 短时间等待
                    
                    # 捕获并处理可能的弹窗
                    page.on("dialog", lambda dialog: dialog.accept())
            
            # 处理下载的文件
            download = await download_info.value
            file_name = f"{query_type}_{uuid.uuid4().hex}.xlsx"
            save_path = os.path.join(self.data_dir, file_name)
            
            # 保存文件并解析
            await download.save_as(save_path)
            processor = ExcelProcessor()
            result = processor.read_file(save_path, query_type)
            return result
        except asyncio.TimeoutError:
            logger.error(f"等待下载超时 ({query_type})")
            return []
        except Exception as e:
            logger.error(f"执行查询 {query_type} 失败: {e}")
            traceback.print_exc()
            return []
        finally:
            # 清理临时文件
            if save_path and os.path.exists(save_path):
                try:
                    os.unlink(save_path)
                except:
                    pass

    async def execute_query(self, query_type, *args, **kwargs):
        """执行查询的统一入口"""
        if not self._initialized:
            await self.initialize()
            
        try:
            # 为每个请求创建独立的上下文和页面
            context = await self.browser.new_context(
                ignore_https_errors=True,
                accept_downloads=True,
                viewport={'width': 1920, 'height': 1080}
            )
            page = await context.new_page()
            page.on("dialog", lambda dialog: dialog.accept())
            page.on("certificateerror", lambda error: error.continue_())
            
            # 确保登录
            if not await self._ensure_login(page):
                return []
                
            # 导航到查询页面
            await page.goto(f'https://10.3.2.201:9943/rntibp/view/complex/{query_type}.html')
            await page.wait_for_load_state('networkidle')
            
            # 准备参数
            if query_type == 'glcx':
                param_dict = {
                    'date_start': args[0],
                    'date_end': args[1],
                    'id_no': args[2]
                }
            elif query_type == 'zzcx':
                param_dict = {
                    'train_date': args[0],
                    'train_code': args[1],
                    'from_station': args[2],
                    'to_station': args[3] if len(args) > 3 else None
                }
            elif query_type == 'plgjcx':
                param_dict = kwargs
            
            # 执行查询
            results = await self._perform_query(page, query_type, param_dict)
            return results
        except Exception as e:
            logger.error(f"执行查询出错: {e}")
            return []
        finally:
            # 清理资源
            if page:
                await page.close()
            if context:
                await context.close()

    # ====================== 独立路由处理 ======================
    async def handle_glcx(self, date_start, date_end, id_no):
        """处理个人查询路由"""
        return await self.execute_query('glcx', date_start, date_end, id_no)

    async def handle_zzcx(self, train_date, train_code, from_station, to_station=None):
        """处理组织查询路由"""
        return await self.execute_query('zzcx', train_date, train_code, from_station, to_station)

    async def handle_plgjcx(self, date_start, date_end, id_no_list):
        """处理批量查询路由"""
        return await self.execute_query(
            'plgjcx',
            date_start=date_start,
            date_end=date_end,
            id_no_list=id_no_list
        )

# ====================== Flask 应用和路由 ======================
def get_automator():
    """获取线程局部的自动化器实例"""
    if not hasattr(_thread_local, 'automator'):
        logger.info("创建新的自动化器实例")
        _thread_local.automator = PlaywrightAutomator()
    return _thread_local.automator

def async_handler(f):
    """将异步视图转换为同步视图的装饰器"""
    @wraps(f)
    def wrapper(*args, **kwargs):
        try:
            return asyncio.run(f(*args, **kwargs))
        except Exception as e:
            logger.error(f"Server Error: {str(e)}")
            traceback.print_exc()
            return make_response(jsonify({'code': 500, 'message': '服务器内部错误'}), 500)
    return wrapper

async def validate_and_execute(data, required_fields, query_method):
    """验证并执行查询逻辑"""
    if not isinstance(data, list):
        return {'code': 400, 'message': '请求数据必须是数组类型'}
    
    # 批量验证所有字段
    for idx, item in enumerate(data):
        missing = [field for field in required_fields if field not in item]
        if missing:
            return {
                'code': 400,
                'message': f'元素{idx}缺少必要字段: {", ".join(missing)}'
            }
    
    try:
        # 异步执行所有任务
        tasks = [query_method(**item) for item in data]
        results = await asyncio.gather(*tasks)
        return {
            'code': 900,
            'data': results
        }
    except Exception as e:
        logger.error(f"执行查询时发生错误: {str(e)}")
        traceback.print_exc()
        return {
            'code': 500,
            'message': f'执行查询时发生错误: {str(e)}'
        }

@app.route('/cyber/test', methods=['GET'])
@async_handler
async def test_endpoint():
    """测试路由 - 验证登录流程"""
    automator = get_automator()
    try:
        title = await automator.test_login()
        return jsonify({
            'code': 900,
            'message': '登录测试成功',
            'title': title
        })
    except Exception as e:
        logger.error(f"测试登录失败: {str(e)}")
        return jsonify({
            'code': 500,
            'message': f'测试登录失败: {str(e)}'
        })

@app.route('/cyber/glcx', methods=['POST'])
@async_handler
async def glcx():
    """个人查询路由"""
    automator = get_automator()
    data = request.get_json()
    result = await validate_and_execute(data, ['date_start', 'date_end', 'id_no'], automator.handle_glcx)
    return jsonify(result)

@app.route('/cyber/zzcx', methods=['POST'])
@async_handler
async def zzcx():
    """组织查询路由"""
    automator = get_automator()
    data = request.get_json()
    result = await validate_and_execute(data, ['train_date', 'train_code', 'from_station', 'to_station'], automator.handle_zzcx)
    return jsonify(result)

@app.route('/cyber/plgjcx', methods=['POST'])
@async_handler
async def plgjcx():
    """批量查询路由"""
    automator = get_automator()
    data = request.get_json()
    
    # 增强参数验证
    if not all(key in data for key in ['date_start', 'date_end']):
        return jsonify({
            'code': 400,
            'message': '请求数据必须包含 date_start 和 date_end 字段'
        })
    
    if not isinstance(data.get('id_no_list'), list):
        return jsonify({
            'code': 400,
            'message': 'id_no_list 必须是数组类型'
        })
    
    # 执行批量查询
    try:
        results = await automator.handle_plgjcx(
            date_start=data['date_start'],
            date_end=data['date_end'],
            id_no_list=data['id_no_list']
        )
        return jsonify({
            'code': 900,
            'data': results
        })
    except Exception as e:
        logger.error(f"批量查询执行失败: {str(e)}")
        traceback.print_exc()
        return jsonify({
            'code': 500,
            'message': f'批量查询执行失败: {str(e)}'
        })

# ====================== 错误处理和清理 ======================
@app.errorhandler(404)
def not_found(e):
    return jsonify({
        'code': 404,
        'message': '请求的资源不存在'
    }), 404

@app.errorhandler(405)
def method_not_allowed(e):
    return jsonify({
        'code': 405,
        'message': '请求方法不允许'
    }), 405

@app.errorhandler(500)
def internal_error(e):
    return jsonify({
        'code': 500,
        'message': '服务器内部错误'
    }), 500

@app.teardown_appcontext
def cleanup(resp_or_exc):
    """应用退出时清理资源"""
    if hasattr(_thread_local, 'automator'):
        automator = _thread_local.automator
        if automator.is_initialized():
            logger.info("清理自动化器资源...")
            loop = asyncio.new_event_loop()
            asyncio.set_event_loop(loop)
            loop.run_until_complete(automator.close())
            loop.close()
        del _thread_local.automator

if __name__ == '__main__':
    app.run(host='0.0.0.0', port=5000, threaded=True)